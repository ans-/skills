#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""langmark.py — mechanical I/O for podcast show language annotation.

Subcommands:
  dump <xlsx>                       Print all rows as JSON. Stdlib only, works anywhere.
  fill <xlsx> <annotations.json>    Write lang_target / lang_from / confidence values and
                                    a live Excel formula in the sql column. Needs openpyxl.
                                    Options: --table pod_shows

The annotations file is a JSON array:
  [{"id": "KoYYtgSKzw", "lang_target": "ja", "lang_from": "", "confidence": 0.78}, ...]

fill refuses to save unless every data row in the sheet has an annotation and every
annotation matches a row — partial writes hide unannotated rows from the reviewer.
"""
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
ANNOT_COLS = ["lang_target", "lang_from", "confidence", "sql"]


def col_letter(idx):
    """1-based index -> column letter."""
    s = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def _cell_ref_to_idx(ref):
    n = 0
    for ch in ref:
        if ch.isalpha():
            n = n * 26 + (ord(ch.upper()) - 64)
    return n - 1


def dump(path):
    """Read sheet1 with stdlib only; formulas come back as '=...' strings."""
    z = zipfile.ZipFile(path)
    shared = []
    if 'xl/sharedStrings.xml' in z.namelist():
        root = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for si in root.findall(NS + 'si'):
            shared.append(''.join(t.text or '' for t in si.iter(NS + 't')))
    root = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
    grid = []
    for row in root.iter(NS + 'row'):
        cells = {}
        for c in row.findall(NS + 'c'):
            idx = _cell_ref_to_idx(c.get('r'))
            f = c.find(NS + 'f')
            if f is not None and f.text:
                cells[idx] = '=' + f.text
                continue
            t = c.get('t')
            v = c.find(NS + 'v')
            if t == 's':
                cells[idx] = shared[int(v.text)] if v is not None else ''
            elif t == 'inlineStr':
                el = c.find(NS + 'is')
                cells[idx] = ''.join(x.text or '' for x in el.iter(NS + 't')) if el is not None else ''
            else:
                cells[idx] = v.text if v is not None else ''
        if cells:
            grid.append([cells.get(i, '') for i in range(max(cells) + 1)])
    if not grid:
        sys.exit("empty sheet")
    header = grid[0]
    rows = []
    for rownum, r in enumerate(grid[1:], start=2):
        d = {'_row': rownum}
        for i, name in enumerate(header):
            if name:
                d[name] = r[i] if i < len(r) else ''
        rows.append(d)
    return {'header': header, 'rows': rows}


def fill(path, annot_path, table):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl missing. Bootstrap once:\n"
                 "  python3 -m venv <scratch>/venv && <scratch>/venv/bin/pip install openpyxl\n"
                 "then rerun with <scratch>/venv/bin/python.")

    raw = json.load(open(annot_path, encoding='utf-8'))
    ann = {}
    for a in raw:
        for lang in (a.get('lang_target', ''), a.get('lang_from', '')):
            if lang and not re.fullmatch(r'[a-z]{2}', lang):
                sys.exit(f"{a['id']}: language code must be 2 lowercase letters or '', got {lang!r}")
        conf = a['confidence']
        if not (0 <= conf <= 1):
            sys.exit(f"{a['id']}: confidence {conf} outside [0, 1]")
        ann[a['id']] = a

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    if 'id' not in col:
        sys.exit(f"no 'id' column in header: {header}")
    for name in ANNOT_COLS:            # append any missing annotation columns
        if name not in col:
            idx = len(header) + 1
            ws.cell(1, idx).value = name
            header.append(name)
            col[name] = idx

    ID, T, F = (col_letter(col[k]) for k in ('id', 'lang_target', 'lang_from'))
    seen, missing = set(), []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, col['id']).value
        if not sid:
            continue
        if sid not in ann:
            missing.append(sid)
            continue
        seen.add(sid)
        a = ann[sid]
        ws.cell(r, col['lang_target']).value = a.get('lang_target', '')
        ws.cell(r, col['lang_from']).value = a.get('lang_from', '')
        c = ws.cell(r, col['confidence'])
        c.value = round(float(a['confidence']), 2)
        c.number_format = '0.00'
        ws.cell(r, col['sql']).value = (
            f'="update {table} set lang_target = \'"&{T}{r}'
            f'&"\', lang_from = \'"&{F}{r}'
            f'&"\', updated_at = CURRENT_TIMESTAMP where id = \'"&{ID}{r}&"\' ;"'
        )
    orphans = set(ann) - seen
    if missing:
        sys.exit(f"rows without annotation, nothing saved: {missing}")
    if orphans:
        sys.exit(f"annotations matching no row, nothing saved: {sorted(orphans)}")

    wb.save(path)
    print(f"filled {len(seen)} rows in {path}")
    print(f"columns: lang_target={T} lang_from={F} confidence={col_letter(col['confidence'])} "
          f"sql={col_letter(col['sql'])} (formula, id={ID})")


def main():
    args = sys.argv[1:]
    if len(args) >= 2 and args[0] == 'dump':
        json.dump(dump(args[1]), sys.stdout, ensure_ascii=False, indent=1)
        print()
    elif len(args) >= 3 and args[0] == 'fill':
        table = 'pod_shows'
        if '--table' in args:
            table = args[args.index('--table') + 1]
        fill(args[1], args[2], table)
    else:
        sys.exit(__doc__)


if __name__ == '__main__':
    main()
